import openpyxl

def grade_score(scores):
  scores = sorted(scores, reverse=True)
  gs = [0.3, 0.7, 1]
  grades = []
  grade_num = [0]
  s_len = len(scores)
  i = 0
  cnt = 0
  for j in range(len(gs)):
    while i < int(s_len * gs[j]) :
      tmp = scores.count(scores[i])
      if cnt + tmp > int(s_len*gs[j]):
        break
      i += tmp
      cnt += tmp
    if(i > 0):
      grade_num.append(i)
      grades.append(grade_pscore(grade_num , scores))
      grades.append(scores[i - 1])
    else:
      grades.append(101)
      grades.append(101)
  return(grades)  
  
def grade_pscore(gr , scores):
  tmp = len(gr) - 1
  g_len = ((gr[tmp] - gr[tmp - 1]) // 2) + gr[tmp - 1]
  i = gr[tmp - 1]
  cnt = gr[tmp - 1]
  while i <= g_len :
    tmp = scores.count(scores[i])
    if cnt + tmp > g_len :
      break
    i += tmp
    cnt += tmp
  if(i > 0):
    return scores[i - 1]
  else:
    return 101
   

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
scores = []

row_id = 1
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column=3).value * 0.3
        sum_v += ws.cell(row = row_id, column=4).value * 0.35
        sum_v += ws.cell(row = row_id, column=5).value * 0.34
        sum_v += ws.cell(row = row_id, column=6).value
        ws.cell(row = row_id, column=7).value = sum_v
        scores.append(sum_v)
    row_id += 1

grades = grade_score(scores)
row_id = 1
for row in ws:
    if row_id != 1:
        if ws.cell(row = row_id, column=7).value >= grades[0]:
            ws.cell(row = row_id, column=8).value = 'A+'
        elif ws.cell(row = row_id, column=7).value >= grades[1]:
            ws.cell(row = row_id, column=8).value = 'A'
        elif ws.cell(row = row_id, column=7).value >= grades[2]:
            ws.cell(row = row_id, column=8).value = 'B+'
        elif ws.cell(row = row_id, column=7).value >= grades[3]:
            ws.cell(row = row_id, column=8).value = 'B'
        elif ws.cell(row = row_id, column=7).value >= grades[4]:
            ws.cell(row = row_id, column=8).value = 'C+'
        else:
             ws.cell(row = row_id, column=8).value = 'C'
    row_id += 1         
wb.save("student.xlsx")             